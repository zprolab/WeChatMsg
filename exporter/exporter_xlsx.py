import os
import time
import traceback

from wxManager import Me, MessageType
from wxManager.decrypt.decrypt_dat import batch_decode_image_multiprocessing
from wxManager.log import logger
from wxManager.model import Message
from exporter.exporter import ExporterBase, copy_files, decode_audios, get_new_filename

from PIL import JpegImagePlugin
from PIL import ImageFile

from PIL import Image as PILImage

from wxManager.parser.link_parser import wx_sport, wx_collection_data, wx_pay_data

JpegImagePlugin._getmp = lambda x: None
ImageFile.LOAD_TRUNCATED_IMAGES = True


def add_hyperlink(doc, row, column, hyperlink):
    from openpyxl.styles import Font
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter
    Image.MAX_IMAGE_PIXELS = None
    cell = doc.cell(row=row, column=column)
    cell.hyperlink = hyperlink
    # 添加样式来改变超链接文本的颜色和下划线
    font = Font(color="0000FF", underline="single")  # 蓝色和单下划线
    cell.font = font


def find_image_with_known_extensions(img_path):
    # 常见的图片后缀名
    extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp']
    directory = os.path.dirname(img_path)
    filename = os.path.basename(img_path)

    for ext in extensions:
        # 构造完整路径
        full_path = os.path.join(directory, f"{filename}{ext}")
        # 检查文件是否存在
        if os.path.isfile(full_path):
            return full_path

    return None


class ExcelExporter(ExporterBase):
    row = 2

    def add_member_info(self, sheet):
        if self.contact.is_chatroom():
            columns = ['wxid', '微信号', '类型', '群昵称', '昵称', '头像地址',
                       '头像原图', '标签', '性别', '个性签名', '国家（地区）', '省份', '城市']
            self.group_contacts = self.database.get_chatroom_members(self.contact.wxid)
            # 写入CSV文件
            sheet.append(columns)
            for wxid, contact in self.group_contacts.items():
                sheet.append(
                    [
                        contact.wxid, contact.alias, contact.flag, contact.remark, contact.nickname,
                        contact.small_head_img_url, contact.big_head_img_url, contact.label_name(),
                        contact.gender, contact.signature, *contact.region
                    ]
                )
        else:
            if self.contact.is_public():
                pass
            else:
                columns = (
                    'wxid', '微信号', '类型', '群昵称', '昵称', '头像地址', '头像原图', '标签', '性别', '电话',
                    '个性签名', '国家（地区）', '省份', '城市')
                # 写入CSV文件
                sheet.append(columns)
                contact = self.contact
                sheet.append(
                    [
                        contact.wxid, contact.alias, contact.flag, contact.remark, contact.nickname,
                        contact.small_head_img_url, contact.big_head_img_url, contact.label_name(),
                        contact.gender, contact.signature, *contact.region
                    ]
                )

    def message_to_list(self, message: Message):
        remark = message.display_name
        nickname = message.display_name
        if self.contact.is_chatroom():
            contact = self.group_contacts.get(message.sender_id)
            if contact:
                remark = contact.remark
                nickname = contact.nickname
        else:
            contact = Me() if message.is_sender else self.contact
            remark = contact.remark
            nickname = contact.nickname
        res = [str(message.server_id), message.type_name(), message.display_name, message.str_time, message.to_text(),
               remark, nickname, 'more']
        return res

    def to_excel(self):
        from openpyxl.styles import Font
        import openpyxl
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        Image.MAX_IMAGE_PIXELS = None
        print(f"【开始导出 XLSX {self.contact.remark}】")
        os.makedirs(self.origin_path, exist_ok=True)
        filename = os.path.join(self.origin_path, f"{self.contact.remark}.xlsx")
        filename = get_new_filename(filename)
        columns = ['消息ID', '类型', '发送人', '时间', '内容', '备注', '昵称', '更多信息']
        messages = self.database.get_messages(self.contact.wxid, time_range=self.time_range)
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.create_sheet("聊天记录", 0)
        member_sheet = new_workbook.create_sheet("成员信息", 1)
        self.add_member_info(member_sheet)
        new_sheet.append(columns)
        num = 1
        total_num = len(messages)
        image_tasks = []
        video_tasks = []
        file_tasks = []
        audio_tasks = []
        image_dir = os.path.join(self.origin_path, 'image')
        video_dir = os.path.join(self.origin_path, 'video')
        audio_dir = os.path.join(self.origin_path, 'voice')
        file_dir = os.path.join(self.origin_path, 'file')
        image_index = {}

        def parser_merged(merged_message):
            for msg in merged_message.messages:
                type_ = msg.type
                if type_ == MessageType.Image:
                    msg.set_file_name()
                    image_tasks.append(
                        (
                            os.path.join(Me().wx_dir, msg.path),
                            os.path.join(image_dir, msg.str_time[:7]),
                            msg.file_name
                        )
                    )
                    image_tasks.append(
                        (
                            os.path.join(Me().wx_dir, msg.thumb_path),
                            os.path.join(image_dir, msg.str_time[:7]),
                            msg.file_name + '_t'
                        )
                    )
                    msg.path = f"./image/{msg.str_time[:7]}/{msg.file_name}"
                    msg.thumb_path = f"./image/{msg.str_time[:7]}/{msg.file_name + '_t'}"
                elif type_ == MessageType.File:
                    origin_file_path = os.path.join(Me().wx_dir, msg.path)
                    file_tasks.append(
                        (
                            origin_file_path,
                            os.path.join(file_dir, msg.str_time[:7]),
                            ''
                        )
                    )
                    msg.path = f'./file/{msg.str_time[:7]}/{os.path.basename(origin_file_path)}'
                elif type_ == MessageType.Video:
                    msg.set_file_name()
                    video_tasks.append(
                        (
                            os.path.join(Me().wx_dir, msg.path),
                            os.path.join(video_dir, msg.str_time[:7]),
                            msg.file_name
                        )
                    )
                    ext = os.path.basename(msg.path).split('.')[-1]
                    msg.path = f'./video/{msg.str_time[:7]}/{msg.file_name}.{ext}'
                elif type_ == MessageType.MergedMessages:
                    parser_merged(msg)

        for index, message in enumerate(messages):
            if not self._is_running:
                break
            if index % 1000 == 0:
                self.update_progress_callback(index / total_num)
            if not self.is_selected(message):
                continue
            try:
                new_sheet.append(self.message_to_list(message))
                self.row += 1
            except:
                logger.error(traceback.format_exc())
                continue
            type_ = message.type
            if type_ == MessageType.Image:
                message.set_file_name()
                image_index[message.server_id] = self.row
                image_tasks.append(
                    (
                        os.path.join(Me().wx_dir, message.path),
                        os.path.join(image_dir, message.str_time[:7]),
                        message.file_name
                    )
                )
                image_tasks.append(
                    (
                        os.path.join(Me().wx_dir, message.thumb_path),
                        os.path.join(image_dir, message.str_time[:7]),
                        message.file_name + '_t'
                    )
                )
                message.path = f"./image/{message.str_time[:7]}/{message.file_name}"
                message.thumb_path = f"./image/{message.str_time[:7]}/{message.file_name + '_t'}"
            elif type_ == MessageType.File:
                origin_file_path = os.path.join(Me().wx_dir, message.path)
                file_tasks.append(
                    (
                        origin_file_path,
                        os.path.join(file_dir, message.str_time[:7]),
                        ''
                    )
                )
                if os.path.isfile(origin_file_path):
                    message.path = f'./file/{message.str_time[:7]}/{os.path.basename(origin_file_path)}'
                    add_hyperlink(new_sheet, self.row, 5, message.path)
            elif type_ == MessageType.Video:
                message.set_file_name()
                video_tasks.append(
                    (
                        os.path.join(Me().wx_dir, message.path),
                        os.path.join(video_dir, message.str_time[:7]),
                        message.file_name
                    )
                )
                ext = os.path.basename(message.path).split('.')[-1]
                message.path = f'./video/{message.str_time[:7]}/{message.file_name}.{ext}'
                add_hyperlink(new_sheet, self.row, 5, message.path)
            elif type_ == MessageType.Audio:
                message.set_file_name()
                audio_tasks.append(
                    (
                        self.database.get_media_buffer(message.server_id),
                        os.path.join(audio_dir, message.str_time[:7]),
                        message.file_name
                    )
                )
                message.path = f'./voice/{message.str_time[:7]}/{message.file_name + ".mp3"}'
                add_hyperlink(new_sheet, self.row, 5, message.path)
            elif type_ == MessageType.MergedMessages:
                parser_merged(message)
        # 使用多进程，导出所有图片
        batch_decode_image_multiprocessing(Me().xor_key, image_tasks)

        # 使用多线程，复制文件、视频到导出文件夹
        copy_files(video_tasks + file_tasks)

        decode_audios(audio_tasks)
        if MessageType.Image in self.message_types:
            for index, message in enumerate(messages):
                if message.type == MessageType.Image:
                    if not self.is_selected(message):
                        continue
                    row = image_index[message.server_id]
                    img_path = find_image_with_known_extensions(os.path.join(self.origin_path, message.path))
                    if not img_path:
                        img_path = find_image_with_known_extensions(os.path.join(self.origin_path, message.thumb_path))
                        if not img_path:
                            continue
                    try:
                        # 打开图片以获取其尺寸
                        with PILImage.open(img_path) as img:
                            width, height = img.size
                        max_height = 500
                        # 计算缩放比例
                        scale = min(1.0, max_height / height)

                        # 缩放后的图片尺寸
                        scaled_width = int(width * scale)
                        scaled_height = int(height * scale)

                        # 插入图片
                        img = Image(img_path)
                        img.width = scaled_width
                        img.height = scaled_height

                        # 计算单元格的坐标
                        cell = f"{get_column_letter(5)}{row}"

                        # 将图片添加到工作表
                        new_sheet.add_image(img, cell)

                        # 设置行高
                        new_sheet.row_dimensions[row].height = scaled_height * 0.75  # 0.75 是像素到 Excel 单位的转换因子
                    except:
                        logger.error(traceback.format_exc())
                        pass
        # 获取列的字母表示（A、B、C...）
        col_letter = get_column_letter(1)
        # 设置整列的单元格格式为文本
        for cell in new_sheet[col_letter]:
            cell.number_format = "@"  # "@" 表示文本格式
        try:
            new_workbook.save(filename)
        except PermissionError:
            filename = '.'.join(filename.split('.')[:-1]) + str(int(time.time())) + '.xlsx'
            new_workbook.save(filename)
        self.update_progress_callback(1)
        self.finish_callback(self.exporter_id)
        print(f"【完成导出 XLSX {self.contact.remark}】")

    def public_to_excel(self):
        from openpyxl.styles import Font
        import openpyxl
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        Image.MAX_IMAGE_PIXELS = None

        print(f"【开始导出 XLSX {self.contact.remark}】")
        os.makedirs(self.origin_path, exist_ok=True)
        filename = os.path.join(self.origin_path, f"{self.contact.remark}.xlsx")
        filename = get_new_filename(filename)
        columns = ['日期', '时间', '标题', '描述', '链接', '更多信息']
        messages = self.database.get_messages(self.contact.wxid, time_range=self.time_range)
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.create_sheet("聊天记录", 0)
        new_sheet.append(columns)
        total_num = len(messages)
        for index, message in enumerate(messages):
            if not self._is_running:
                break
            if index % 1000 == 0:
                self.update_progress_callback(index / total_num)
            if not message.type in {MessageType.LinkMessage}:
                continue
            try:
                new_sheet.append([*message.str_time.split(' '), message.title, message.description, message.href])
            except:
                logger.error(traceback.format_exc())
                continue
        # 获取列的字母表示（A、B、C...）
        col_letter = get_column_letter(1)
        # 设置整列的单元格格式为文本
        for cell in new_sheet[col_letter]:
            cell.number_format = "@"  # "@" 表示文本格式
        try:
            new_workbook.save(filename)
        except PermissionError:
            filename = '.'.join(filename.split('.')[:-1]) + str(int(time.time())) + '.xlsx'
            new_workbook.save(filename)
        self.update_progress_callback(1)
        self.finish_callback(self.exporter_id)
        print(f"【完成导出 XLSX {self.contact.remark}】")

    def wx_pay(self):
        from openpyxl.styles import Font
        import openpyxl
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        Image.MAX_IMAGE_PIXELS = None
        print(f"【开始导出 XLSX {self.contact.remark}】")
        os.makedirs(self.origin_path, exist_ok=True)
        filename = os.path.join(self.origin_path, f"{self.contact.remark}.xlsx")
        filename = get_new_filename(filename)
        columns = ['类型', '收款单位', '日期', '时间', '金额', '付款方式', '收单机构', '更多信息']
        messages = self.database.get_messages(self.contact.wxid, time_range=self.time_range)
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.create_sheet("聊天记录", 0)
        new_sheet.append(columns)
        total_num = len(messages)
        for index, message in enumerate(messages):
            if not self._is_running:
                break
            if index % 1000 == 0:
                self.update_progress_callback(index / total_num)
            if not message.type in {MessageType.LinkMessage}:
                continue
            try:
                card_data = wx_pay_data(message.xml_content)
                date, str_time = message.str_time.split(' ')
                if card_data.get('title') in {'记账日报', '「先享后付」服务使用通知', '转入零钱通，五一享收益',
                                              '转入零钱通，端午享收益', '智能手表支付服务已启用', '优惠券领取提醒',
                                              '清明假期收益规则', '「先享后付」服务完成通知', '礼包领取提醒',
                                              '五一假期收益规则提醒', '端午节假期收益规则', '中秋节假期收益规则',
                                              '元旦假期收益规则', '春节假期收益规则', '五一假期收益规则',
                                              '中秋及国庆假期收益规则', '春节赚收益攻略', '「先享后付」服务取消通知',
                                              '揭开骗局，远离诈骗'}:
                    continue
                new_sheet.append(
                    [
                        card_data.get('title'), card_data.get('display_name'), date, str_time,
                        card_data.get('money'), card_data.get('payment_type'), card_data.get('acquiring_institution'),
                        card_data.get('more')
                    ]
                )
            except:
                logger.error(traceback.format_exc())
                continue
        # 获取列的字母表示（A、B、C...）
        col_letter = get_column_letter(1)
        # 设置整列的单元格格式为文本
        for cell in new_sheet[col_letter]:
            cell.number_format = "@"  # "@" 表示文本格式
        try:
            new_workbook.save(filename)
        except PermissionError:
            filename = '.'.join(filename.split('.')[:-1]) + str(int(time.time())) + '.xlsx'
            new_workbook.save(filename)
        self.update_progress_callback(1)
        self.finish_callback(self.exporter_id)
        print(f"【完成导出 XLSX {self.contact.remark}】")

    def wx_collect(self):
        from openpyxl.styles import Font
        import openpyxl
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        Image.MAX_IMAGE_PIXELS = None

        print(f"【开始导出 XLSX {self.contact.remark}】")
        os.makedirs(self.origin_path, exist_ok=True)
        filename = os.path.join(self.origin_path, f"{self.contact.remark}.xlsx")
        filename = get_new_filename(filename)
        columns = ['类型', '日期', '时间', '金额', '详细信息', '汇总', '备注', '更多信息']
        messages = self.database.get_messages(self.contact.wxid, time_range=self.time_range)
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.create_sheet("聊天记录", 0)
        new_sheet.append(columns)
        total_num = len(messages)
        for index, message in enumerate(messages):
            if not self._is_running:
                break
            if index % 1000 == 0:
                self.update_progress_callback(index / total_num)
            if not message.type in {MessageType.LinkMessage}:
                continue
            try:
                card_data = wx_collection_data(message.xml_content)
                date, str_time = message.str_time.split(' ')
                new_sheet.append(
                    [
                        card_data.get('title'), date, str_time, card_data.get('money'), card_data.get('display_name'),
                        card_data.get('summary'), card_data.get('more')
                    ]
                )
            except:
                logger.error(traceback.format_exc())
                continue
        # 获取列的字母表示（A、B、C...）
        col_letter = get_column_letter(1)
        # 设置整列的单元格格式为文本
        for cell in new_sheet[col_letter]:
            cell.number_format = "@"  # "@" 表示文本格式
        try:
            new_workbook.save(filename)
        except PermissionError:
            filename = '.'.join(filename.split('.')[:-1]) + str(int(time.time())) + '.xlsx'
            new_workbook.save(filename)
        self.update_progress_callback(1)
        self.finish_callback(self.exporter_id)
        print(f"【完成导出 XLSX {self.contact.remark}】")

    def wx_sport(self):
        from openpyxl.styles import Font
        import openpyxl
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
        Image.MAX_IMAGE_PIXELS = None


        print(f"【开始导出 XLSX {self.contact.remark}】")
        os.makedirs(self.origin_path, exist_ok=True)
        filename = os.path.join(self.origin_path, f"{self.contact.remark}.xlsx")
        filename = get_new_filename(filename)
        columns = ['日期', '排名', '步数', '当日冠军', '当日冠军步数', '更多信息']
        messages = self.database.get_messages(self.contact.wxid, time_range=self.time_range)
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.create_sheet("聊天记录", 0)
        new_sheet.append(columns)
        total_num = len(messages)
        for index, message in enumerate(messages):
            if not self._is_running:
                break
            if index and index % 1000 == 0:
                self.update_progress_callback(index / total_num)
            if not message.type in {MessageType.LinkMessage}:
                continue
            try:
                card_data = wx_sport(message.xml_content)
                champion_name = ''
                if not card_data.get('rank_list'):
                    champion = {}
                else:
                    champion = card_data.get('rank_list')[0]
                    contact = self.database.get_contact_by_username(champion.get('username'))
                    champion_name = contact.remark
                new_sheet.append(
                    [
                        message.str_time.split(' ')[0], card_data.get('rank'), card_data.get('score'),
                        champion_name, champion.get('score')
                    ]
                )
            except:
                logger.error(traceback.format_exc())
                continue
        # 获取列的字母表示（A、B、C...）
        col_letter = get_column_letter(1)
        # 设置整列的单元格格式为文本
        for cell in new_sheet[col_letter]:
            cell.number_format = "@"  # "@" 表示文本格式
        try:
            new_workbook.save(filename)
        except PermissionError:
            filename = '.'.join(filename.split('.')[:-1]) + str(int(time.time())) + '.xlsx'
            new_workbook.save(filename)
        self.update_progress_callback(1)
        self.finish_callback(self.exporter_id)
        print(f"【完成导出 XLSX {self.contact.remark}】")

    def run(self):
        if self.contact.is_public():
            if self.contact.wxid == 'gh_3dfda90e39d6':
                self.wx_pay()
            elif self.contact.wxid == 'gh_f0a92aa7146c':
                self.wx_collect()
            elif self.contact.wxid == 'gh_43f2581f6fd6':
                self.wx_sport()
            else:
                self.public_to_excel()
        else:
            self.to_excel()
